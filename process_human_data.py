# This document parses the metadata, current, and voltage traces from all NWB files in a given directory
import os
import pandas as pd
import numpy as np
import ast
import gc
from datetime import datetime
from pynwb import NWBHDF5IO
from matplotlib import pyplot as plt
import re
from openpyxl import load_workbook
from pathlib import Path

# Set to True to enable verbose/debug output in terminal
VERBOSE = False

# -------------------------------------------------------------
# Checkpoint function
# -------------------------------------------------------------

def checkpoint(milestone_name: str) -> bool:
    """
    Auto-yes checkpoint - no longer prompts user.
    
    Args:
        milestone_name: Description of the completed milestone
        
    Returns:
        True (always proceeds)
    """
    if VERBOSE:
        print("\n" + "="*70)
        print(f"✓ {milestone_name} was successful!")
        print("="*70)
    return True


# -------------------------------------------------------------
# Parsing helpers
# -------------------------------------------------------------

def safe_getattr(obj, attr, default=None):
    val = getattr(obj, attr, default)
    try:
        if val is None:
            return default
        if hasattr(val, "__iter__") and not isinstance(val, str):
            return list(val)
        else:
            return val
    except Exception:
        return default


def normalize_age(age_str):
    """
    Normalize age to years format (e.g., P62.0Y).
    Converts from days (P21170.0D) to years if needed.
    
    Args:
        age_str: Age string in ISO 8601 duration format (e.g., "P21170.0D" or "P62.0Y")
    
    Returns:
        Age string in years format (e.g., "P62.0Y"), or original if already in years or unparseable
    """
    if not age_str or not isinstance(age_str, str):
        return age_str
    
    # Already in years format
    if 'Y' in age_str:
        return age_str
    
    # Convert from days to years
    if 'D' in age_str:
        try:
            # Extract numeric value from format like "P21170.0D"
            days_str = age_str.replace('P', '').replace('D', '')
            days = float(days_str)
            years = days / 365.25  # Account for leap years
            return f"P{years:.1f}Y"
        except (ValueError, AttributeError):
            return age_str
    
    return age_str


def parse_keywords(kw):
    # numpy array
    if isinstance(kw, np.ndarray):
        return kw.tolist()

    # python list or tuple (NWB keywords often tuples!)
    if isinstance(kw, (list, tuple)):
        return list(kw)

    # string representation of a list
    if isinstance(kw, str):
        try:
            return list(ast.literal_eval(kw))
        except Exception:
            return [kw]

    # fallback
    return []


def rec_goal(desc):
    required = [
        "Whole-cell patch-clamp",
        "current-clamp",
        "I-V protocol",
        "Human cortex L2/3",
        "pyramidalcell"
    ]
    return "Intrinsic properties" if all(x in desc for x in required) else ""


def parse_cell_depth(kw): 
    return "Layer 2/3" if any("Layer 2/3" in k for k in kw) else ""


def parse_cell_type(kw):
    return "pyramidalcell" if any("pyramidalcell" in k.lower() for k in kw) else ""


def parse_brain_loc(kw):
    return "Human Neocortex" if any("Human Neocortex" in k for k in kw) else ""


def get_sample_rate_hz(nwb):
    """
    Return the sampling rate (Hz) from the first voltage trace.
    """
    for name, ts in nwb.acquisition.items():
        if hasattr(ts, "rate") and ts.rate is not None:
            return float(ts.rate) 
        else:
            print("⚠ No rate found in trace:", name)
    return None

# ts = nwb.acquisition[name]

#         rate = getattr(ts, "rate", None)

def get_time_vector(ts):
    if ts.timestamps is not None:
        return np.asarray(ts.timestamps)
    elif ts.rate is not None:
        return np.arange(len(ts.data)) / ts.rate + ts.starting_time
    else:
        raise ValueError("Cannot determine time vector for TimeSeries")

# def get_time_vector(ts):
#     if hasattr(ts, "timestamps") and ts.timestamps is not None:
#         return np.array(ts.timestamps[:])
#     elif hasattr(ts, "starting_time") and hasattr(ts, "rate") and ts.rate is not None:
#         return np.arange(len(ts.data)) / ts.rate + ts.starting_time
#     else:
#         return np.arange(len(ts.data))


def convert_voltage(data, unit):
    """Convert voltage to mV if needed."""
    if unit.lower() == "v" or unit.lower() == "volts":
        return data * 1e3, "mV"
    return data, unit


def convert_current(data, unit):
    """Convert current to pA if needed."""
    if unit.lower() == "a" or unit.lower() == "amperes":
        return data * 1e12, "pA"
    return data, unit

def sweep_sort_key(name):
    """
    Extract trailing integer from sweep name for natural sorting.
    Falls back to name if no number found.
    """
    match = re.search(r'(\d+)$', name)
    return int(match.group(1)) if match else name

def extract_sweep_number(name):
    match = re.search(r'(\d+)$', name)
    if match:
        return int(match.group(1))
    else:
        raise ValueError(f"Cannot extract numeric sweep from name: {name}")

def extract_current_voltage_from_nwb(nwb, out_dir, cellNum, plot=False):
    os.makedirs(out_dir, exist_ok=True)

    # --------------------------
    # Discover traces
    # --------------------------
    voltage_traces = [
        k for k in nwb.acquisition.keys()
        if "vm" in k.lower() or "volt" in k.lower()
    ]
    voltage_traces = sorted(voltage_traces, key=sweep_sort_key)

    current_traces = list(nwb.stimulus.keys()) if nwb.stimulus else []
    current_traces = sorted(current_traces, key=sweep_sort_key)

    if VERBOSE:
        print("Voltage traces:", voltage_traces)
        print("Current traces:", current_traces)

    voltage_rows = []
    current_rows = []
    
    # Collect plot data if plotting is enabled
    voltage_plot_data = [] if plot else None
    current_plot_data = [] if plot else None

    # --------------------------
    # Voltage extraction
    # --------------------------
    for name in voltage_traces:
        ts = nwb.acquisition[name]

        t = get_time_vector(ts)
        data = np.asarray(ts.data)

        unit = getattr(ts, "unit", None)
        data, unit = convert_voltage(data, unit)

        rate = getattr(ts, "rate", None)

        sweep_num = extract_sweep_number(name)

        df = pd.DataFrame({
            "sweep": sweep_num,
            "t_s": t,
            "value": data,
            "unit": unit
        })

        voltage_rows.append(df)

        if plot:
            voltage_plot_data.append({'sweep': sweep_num, 'name': name, 't': t, 'd': data, 'unit': unit})


    # --------------------------
    # Current extraction
    # --------------------------
    for name in current_traces:
        ts = nwb.stimulus[name]

        t = get_time_vector(ts)
        data = np.asarray(ts.data)

        unit = getattr(ts, "unit", None)
        data, unit = convert_current(data, unit)

        rate = getattr(ts, "rate", None)

        sweep_num = extract_sweep_number(name)

        df = pd.DataFrame({
            "sweep": sweep_num,
            "t_s": t,
            "value": data,
            "unit": unit
        })

        current_rows.append(df)

        if plot:
            current_plot_data.append({'sweep': sweep_num, 'name': name, 't': t, 'd': data, 'unit': unit})

    # --------------------------
    # Generate grid plots
    # --------------------------
    if plot and voltage_plot_data:
        n_sweeps = len(voltage_plot_data)
        n_cols = min(5, n_sweeps)
        n_rows = int(np.ceil(n_sweeps / n_cols))
        
        fig, axes = plt.subplots(n_rows, n_cols, figsize=(n_cols * 4, n_rows * 3))
        if n_sweeps == 1:
            axes = np.array([axes])
        axes = axes.flatten()
        
        for i, item in enumerate(voltage_plot_data):
            ax = axes[i]
            ax.plot(item['t'], item['d'], lw=0.8)
            ax.set_xlabel("Time (s)")
            ax.set_ylabel(item['unit'] if item['unit'] else "value")
            ax.set_title(f"Voltage: {item['name']}")
        
        # Hide unused subplots
        for j in range(n_sweeps, len(axes)):
            axes[j].axis('off')
        
        fig.suptitle(f'Voltage Traces - {cellNum}', fontsize=16, fontweight='bold', y=0.995)
        plt.tight_layout()
        voltage_grid_path = os.path.join(out_dir, "voltage_grid.png")
        plt.savefig(voltage_grid_path, dpi=300)
        plt.close()
        print(f"✔ Saved voltage grid → {voltage_grid_path}")
    
    if plot and current_plot_data:
        n_sweeps = len(current_plot_data)
        n_cols = min(5, n_sweeps)
        n_rows = int(np.ceil(n_sweeps / n_cols))
        
        fig, axes = plt.subplots(n_rows, n_cols, figsize=(n_cols * 4, n_rows * 3))
        if n_sweeps == 1:
            axes = np.array([axes])
        axes = axes.flatten()
        
        for i, item in enumerate(current_plot_data):
            ax = axes[i]
            ax.plot(item['t'], item['d'], lw=0.8)
            ax.set_xlabel("Time (s)")
            ax.set_ylabel(item['unit'] if item['unit'] else "value")
            ax.set_title(f"Current: {item['name']}")
        
        # Hide unused subplots
        for j in range(n_sweeps, len(axes)):
            axes[j].axis('off')
        
        fig.suptitle(f'Current Traces - {cellNum}', fontsize=16, fontweight='bold', y=0.995)
        plt.tight_layout()
        current_grid_path = os.path.join(out_dir, "current_grid.png")
        plt.savefig(current_grid_path, dpi=300)
        plt.close()
        print(f"✔ Saved current grid → {current_grid_path}")

    # --------------------------
    # Save CSVs
    # --------------------------
    if voltage_rows:
        voltage_all = pd.concat(voltage_rows, ignore_index=True)
        voltage_csv = os.path.join(out_dir, f"mV_{cellNum}.csv")
        voltage_parquet = os.path.join(out_dir, f"mV_{cellNum}.parquet")
        assert voltage_all["sweep"].dtype.kind in "iu"
        # voltage_all.to_csv(voltage_csv, index=False)  # CSV creation disabled, using parquet only
        voltage_all.to_parquet(voltage_parquet, index=False)
        print(f"✔ Saved voltage traces → {voltage_parquet}")
        # Clear from memory
        del voltage_all

    if current_rows:
        current_all = pd.concat(current_rows, ignore_index=True)
        current_csv = os.path.join(out_dir, f"pA_{cellNum}.csv")
        current_parquet = os.path.join(out_dir, f"pA_{cellNum}.parquet")
        assert current_all["sweep"].dtype.kind in "iu"
        # current_all.to_csv(current_csv, index=False)  # CSV creation disabled, using parquet only
        current_all.to_parquet(current_parquet, index=False)
        print(f"✔ Saved current traces → {current_parquet}")
        # Clear from memory
        del current_all
    
    # Clear plot data and row lists
    if voltage_rows:
        del voltage_rows
    if current_rows:
        del current_rows
    
    # Close all matplotlib figures
    plt.close('all')


# -------------------------------------------------------------
# MAIN LOOP
# -------------------------------------------------------------

if __name__ == "__main__":
    import sys
    
    # Support command-line arguments from main.py
    if len(sys.argv) > 1:
        # Called from main.py with arguments
        parent_dir = Path(sys.argv[1]).expanduser()
        log_output_dir = Path(sys.argv[2]).expanduser()
        mice_metadata_excel = Path(sys.argv[3]).expanduser()
        
        # Template validation already done in main.py - just use the path
    else:
        # Interactive mode (standalone use)
        parent_dir = Path(input("Parent directory for all subjects: ")).expanduser()
        log_output_dir = Path(input("Output directory for excel metadata: ")).expanduser()
        mice_metadata_excel = Path(input("Path to mice excel metadata: ")).expanduser()

    # Validate parent directory
    if not parent_dir.is_dir():
        raise FileNotFoundError(f"Parent directory not found: {parent_dir}")

    # Prepare output directory
    log_output_dir.mkdir(parents=True, exist_ok=True)
    log_output_path = log_output_dir / "human_ephys_log.xlsx"

    # Validate Excel template exists
    if not mice_metadata_excel.is_file():
        raise FileNotFoundError(f"Metadata Excel file not found: {mice_metadata_excel}")

    if mice_metadata_excel.suffix not in {".xlsx", ".xlsm"}:
        raise ValueError("Metadata file must be .xlsx or .xlsm")

    try:
        wb = load_workbook(mice_metadata_excel)
        ws = wb.active
    except Exception as e:
        raise RuntimeError(f"Failed to load Excel file: {e}")

    HEADER_ROW = 3
    column_names = []
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=HEADER_ROW, column=col).value
        if val is None:
            break
        column_names.append(val)

    column_names.append("brain_loc")
    # Note: sampleRate_Hz is no longer added to Excel

    all_rows = [] 

    for subfolder in os.listdir(parent_dir): 
        nwb_subfolder_path = os.path.join(parent_dir, subfolder) 
        if not os.path.isdir(nwb_subfolder_path): 
            continue # skip files at parent level 
        print(f"\n📁 Entering folder: {subfolder}") 
        # nwb_folder = "/Users/snehajaikumar/Human Data to share/sub-131008" 
        # folder containing NWB and parquet files 
        for root, _, files in os.walk(nwb_subfolder_path): 
            cell_count = 0 
            for fname in files: 
                if fname.endswith(".nwb"): 
                    print(f"\nProcessing NWB file: {fname}") 
                    cell_count += 1 
                    nwb_path = os.path.join(root, fname) 
                    # Open the NWB file using pynwb 
                    try: 
                        # with NWBHDF5IO(nwb_path, "r") as io: 
                        io = NWBHDF5IO(nwb_path, "r") 
                        nwb = io.read() 
                    except Exception as e: 
                        print(f"⚠ Error processing {fname}: {e}") 
                        continue 
                    
                    # Check if this is a single protocol file (skip mixed protocol files)
                    protocols = set()
                    sampling_rates = set()
                    if nwb.stimulus:
                        for stim_name in nwb.stimulus.keys():
                            stim_series = nwb.stimulus[stim_name]
                            series_type = getattr(stim_series, 'neurodata_type', '')
                            if 'VoltageClamp' in series_type or 'voltage_clamp' in stim_name.lower():
                                protocols.add('VoltageClamp')
                            elif 'CurrentClamp' in series_type or 'current_clamp' in stim_name.lower():
                                protocols.add('CurrentClamp')
                            
                            # Collect sampling rate
                            if hasattr(stim_series, 'rate') and stim_series.rate:
                                sampling_rates.add(float(stim_series.rate))
                    
                    # Skip if mixed protocol (only process single protocol files)
                    if len(protocols) > 1:
                        print(f"  ⊘ Skipping: Mixed protocol file (not single)")
                        io.close()
                        continue
                    
                    # Determine sampling rate (should be single value for single protocol)
                    if len(sampling_rates) == 1:
                        sample_rate = list(sampling_rates)[0]
                    elif len(sampling_rates) > 1:
                        # Unexpected: multiple rates in single protocol file
                        print(f"⚠ WARNING: Found multiple sampling rates in single protocol: {sampling_rates}")
                        sample_rate = sorted(list(sampling_rates), reverse=True)
                    else:
                        # Fallback if no rate detected
                        sample_rate = 50000.0
                    
                    # -------------------------- # Extract metadata properly # ------------------------- 
                    keywords = list(safe_getattr(nwb, "keywords", [])) 
                    session_description = safe_getattr(nwb, "session_description", "") 
                    identifier = safe_getattr(nwb, "identifier", "") 
                    # use recording date 
                    file_create_date = safe_getattr(nwb, "nwb.session_start_time", None) 
                    lab = safe_getattr(nwb, "lab", "") 
                    institution = safe_getattr(nwb, "institution", "") 
                    experiment_description = safe_getattr(nwb, "experiment_description", "") 
                    experimenter = safe_getattr(nwb, "experimenter", []) 
                    subject = nwb.subject 
                    subject_id = safe_getattr(subject, "subject_id", "") 
                    age = safe_getattr(subject, "age", "")
                    age = normalize_age(age)  # Convert days to years if needed
                    sex = safe_getattr(subject, "sex", "") 
                    dt = safe_getattr(nwb, "session_start_time", None) 
                    genotype = safe_getattr(subject, "genotype", "")
                    
                    # Use institution as fallback for facility if lab is not available
                    facility = lab if lab else institution
                    
                    # -------------------------- # Build row for Excel # -------------------------- 
                    row = { 
                        "recDate": dt.strftime("%Y%m%d") if dt else "", 
                        "facility": facility, 
                        "genotype": genotype, 
                        "sex": sex, 
                        "age_at_recDate": age, 
                        "recGoal": rec_goal(experiment_description), 
                        "cellNum": str(subject_id) + "_" +str(cell_count), 
                        "fileNum": identifier, 
                        "expInfoTime": dt.isoformat() if dt else "", 
                        "cellDepth": parse_cell_depth(keywords), 
                        "cellType": parse_cell_type(keywords), 
                        "remarks": subject_id, 
                        "brain_loc": parse_brain_loc(keywords), 
                    } 

                    full_row = {col: row.get(col, "") for col in column_names} 

                    # Note: sampleRate_Hz is no longer added to Excel
                    all_rows.append(full_row) 

                    # -------------------------- # Extract CURRENT + VOLTAGE # -------------------------- 
                    base_name = os.path.splitext(fname)[0] 
                    cell_folder = os.path.join(nwb_subfolder_path, base_name) 
                    os.makedirs(cell_folder, exist_ok=True) 
                    try: 
                        extract_current_voltage_from_nwb( 
                            nwb, 
                            cell_folder, 
                            row["cellNum"], plot=True # set True if you want plots 
                        ) 
                    except Exception as e: 
                        print(f"⚠ Failed on {fname}: {e}") 
                        
                    # -------------------------- # Create manifest with metadata # -------------------------- 
                    manifest_path = os.path.join(cell_folder, "manifest.json") 
                    
                    # Add sampling rate to metadata
                    full_row['sampleRate_Hz'] = sample_rate
                    
                    manifest = { 
                        "file_id": identifier, 
                        "nwb_path": os.path.abspath(nwb_path), 
                        "tables": { 
                            "mv": f"mV_{row['cellNum']}.parquet", 
                            "pa": f"pA_{row['cellNum']}.parquet" 
                        }, 
                        "meta": full_row 
                    } 
                    
                    with open(manifest_path, "w") as f: 
                        import json 
                        json.dump(manifest, f, indent=2, default=str) 
                    print(f"✔ Saved manifest → {manifest_path}") 
                    io.close()
                    
                    # Force garbage collection after each file to free memory
                    gc.collect()
    # ------------------------------------------------------------- # WRITE FINAL COMBINED LOG # ------------------------------------------------------------- 
    df = pd.DataFrame(all_rows, columns=column_names)
    
    # Check if file exists and append if it does
    if log_output_path.exists():
        existing_df = pd.read_excel(log_output_path)
        
        # Ensure columns match
        for col in df.columns:
            if col not in existing_df.columns:
                existing_df[col] = ""
        for col in existing_df.columns:
            if col not in df.columns:
                df[col] = ""
        
        # Append new rows
        combined_df = pd.concat([existing_df, df], ignore_index=True)
        combined_df.to_excel(log_output_path, index=False)
        print("\n=====================================")
        print("✔ Appended to human_ephys_log!")
        print(f"✔ Saved to {log_output_path}")
        print(f"✔ New rows added: {len(df)}")
        print(f"✔ Total rows now: {len(combined_df)}")
        print("=====================================\n")
    else:
        df.to_excel(log_output_path, index=False)
        print("\n=====================================")
        print("✔ Completed human_ephys_log!")
        print(f"✔ Saved to {log_output_path}")
        print(f"✔ Total rows: {len(df)}")
        print("=====================================\n")

    # MILESTONE CHECKPOINT: Bundle creation and metadata spreadsheet
    if not checkpoint("Bundle directories created with parquet files and human_ephys_log.xlsx metadata spreadsheet"):
        print("\n⚠ User chose not to proceed. Bundles have been created but analysis will not be run.")
        import sys
        sys.exit(0)

    