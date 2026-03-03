#!/usr/bin/env python3
"""
Generic Mixed Protocol NWB Extractor
====================================

This script extracts data from NWB files that contain mixed protocols
(e.g., both VoltageClamp and CurrentClamp recordings).

Features:
  • Detects protocol type per sweep (VoltageClamp vs CurrentClamp)
  • Stores per-sweep sampling rates in rate_Hz column
  • Creates enhanced manifest.json with protocol metadata
  • Works with any mixed protocol NWB file automatically

Usage:
  $ python process_human_data_mixed_protocol.py
  
  Then follow prompts to enter:
    - NWB file path
    - Output directory (optional, auto-generated)
    - Cell ID (optional, extracted from filename)
"""

import os
import pandas as pd
import numpy as np
import re
import json
import gc
from datetime import datetime
from pynwb import NWBHDF5IO
from openpyxl import load_workbook
from pathlib import Path
from matplotlib import pyplot as plt

# Set to True to enable verbose/debug output in terminal
VERBOSE = False


def get_time_vector(ts):
    """Extract time vector from timeseries object."""
    if hasattr(ts, "timestamps") and ts.timestamps is not None:
        return np.asarray(ts.timestamps)
    elif hasattr(ts, "rate") and ts.rate is not None:
        starting_time = getattr(ts, "starting_time", 0.0)
        return np.arange(len(ts.data)) / ts.rate + starting_time
    raise ValueError("Cannot get time vector")


def safe_getattr(obj, attr, default=None):
    """Safely get attribute from object, handling iterables and None values."""
    val = getattr(obj, attr, default)
    try:
        if val is None:
            return default
        if hasattr(val, "__iter__") and not isinstance(val, str):
            return list(val)
        else:
            return val
    except Exception:
        return default


def normalize_age(age_str):
    """
    Normalize age to years format (e.g., P62.0Y).
    Converts from days (P21170.0D) to years if needed.
    
    Args:
        age_str: Age string in ISO 8601 duration format (e.g., "P21170.0D" or "P62.0Y")
    
    Returns:
        Age string in years format (e.g., "P62.0Y"), or original if already in years or unparseable
    """
    if not age_str or not isinstance(age_str, str):
        return age_str
    
    # Already in years format
    if 'Y' in age_str:
        return age_str
    
    # Convert from days to years
    if 'D' in age_str:
        try:
            # Extract numeric value from format like "P21170.0D"
            days_str = age_str.replace('P', '').replace('D', '')
            days = float(days_str)
            years = days / 365.25  # Account for leap years
            return f"P{years:.1f}Y"
        except (ValueError, AttributeError):
            return age_str
    
    return age_str


def extract_sweep_number(name):
    """Extract sweep number from series name."""
    m = re.search(r'data_(\d+)', name)
    return int(m.group(1)) if m else int(name.split('_')[0])


def get_protocol_type(stim_series):
    """Determine protocol type (VoltageClamp or CurrentClamp) from stimulus series."""
    series_type = type(stim_series).__name__
    if 'VoltageClamp' in series_type:
        return 'VoltageClamp'
    elif 'CurrentClamp' in series_type:
        return 'CurrentClamp'
    return 'Unknown'


def parse_cell_depth(keywords):
    """Extract cell depth from keywords."""
    return "Layer 2/3" if any("Layer 2/3" in k for k in keywords) else ""


def parse_cell_type(keywords):
    """Extract cell type from keywords."""
    return "pyramidalcell" if any("pyramidalcell" in k.lower() for k in keywords) else ""


def extract_from_mixed_protocol_nwb(nwb_path, out_dir, cellNum, plot=False):
    """
    Extract data from mixed protocol NWB file.
    
    Args:
        nwb_path: Path to NWB file
        out_dir: Output directory for extracted data
        cellNum: Cell number in format "subjectID_cellCount" (e.g., "sub-1000610030_1")
        plot: Whether to generate plots
        
    Returns:
        tuple: (sampling_rates_set, protocol_info_dict) for manifest creation
    """
    os.makedirs(out_dir, exist_ok=True)
    
    # Get filename for display
    nwb_filename = Path(nwb_path).name
    
    print(f"\n{'='*70}")
    print(f"MIXED PROTOCOL EXTRACTION")
    print(f"File: {nwb_filename}")
    print(f"{'='*70}\n")
    
    with NWBHDF5IO(nwb_path, 'r') as io:
        nwb = io.read()
        
        acq = {extract_sweep_number(n): (n, nwb.acquisition[n]) for n in nwb.acquisition.keys()}
        stim = {extract_sweep_number(n): (n, nwb.stimulus[n]) for n in nwb.stimulus.keys()}
        
        stim_rows, resp_rows = [], []
        protocol_info = {}  # Track protocol type and rate for each sweep
        
        # Collect plot data if plotting is enabled
        stim_plot_data = [] if plot else None
        resp_plot_data = [] if plot else None
        
        # Extract stimulus data for each sweep
        if VERBOSE:
            print("STIMULUS EXTRACTION")
            print("-"*70)
        for sw in sorted(stim.keys()):
            n, c = stim[sw]
            protocol = get_protocol_type(c)
            rate = getattr(c, 'rate', 50000.0)
            protocol_info[sw] = {'protocol': protocol, 'rate': rate}
            
            t, d, r = get_time_vector(c), np.asarray(c.data, dtype=np.float32), rate
            stim_rows.append(pd.DataFrame({
                'sweep': sw,
                't_s': t,
                'value': d,
                'unit': c.unit.lower()
            }))
            
            if plot:
                stim_plot_data.append({'sweep': sw, 'name': n, 't': t, 'd': d})
        
        # Extract response data for each sweep
        if VERBOSE:
            print("RESPONSE EXTRACTION")
            print("-"*70)
        for sw in sorted(acq.keys()):
            n, c = acq[sw]
            t, d, r = get_time_vector(c), np.asarray(c.data, dtype=np.float32), getattr(c, 'rate', None)
            resp_rows.append(pd.DataFrame({
                'sweep': sw,
                't_s': t,
                'value': d,
                'unit': c.unit.lower()
            }))
            
            if plot:
                resp_plot_data.append({'sweep': sw, 'name': n, 't': t, 'd': d})
        
        # Create grid plots if plotting is enabled
        if plot and stim_plot_data:
            n_sweeps = len(stim_plot_data)
            n_cols = min(5, n_sweeps)
            n_rows = int(np.ceil(n_sweeps / n_cols))
            
            fig, axes = plt.subplots(n_rows, n_cols, figsize=(n_cols * 4, n_rows * 3))
            if n_sweeps == 1:
                axes = np.array([axes])
            axes = axes.flatten()
            
            for idx, sweep_data in enumerate(stim_plot_data):
                axes[idx].plot(sweep_data['t'], sweep_data['d'], lw=0.8)
                axes[idx].set_title(f"Sweep {sweep_data['sweep']}")
                axes[idx].set_xlabel('Time (s)')
                axes[idx].set_ylabel('Stimulus')
            
            # Hide unused subplots
            for idx in range(n_sweeps, len(axes)):
                axes[idx].axis('off')
            
            fig.suptitle(f'Stimulus Traces - {cellNum}', fontsize=16, fontweight='bold', y=0.995)
            plt.tight_layout()
            plt.savefig(os.path.join(out_dir, "stimulus_grid.png"), dpi=300)
            plt.close()
            print(f"✔ Saved stimulus grid plot → stimulus_grid.png")
        
        if plot and resp_plot_data:
            n_sweeps = len(resp_plot_data)
            n_cols = min(5, n_sweeps)
            n_rows = int(np.ceil(n_sweeps / n_cols))
            
            fig, axes = plt.subplots(n_rows, n_cols, figsize=(n_cols * 4, n_rows * 3))
            if n_sweeps == 1:
                axes = np.array([axes])
            axes = axes.flatten()
            
            for idx, sweep_data in enumerate(resp_plot_data):
                axes[idx].plot(sweep_data['t'], sweep_data['d'], lw=0.8)
                axes[idx].set_title(f"Sweep {sweep_data['sweep']}")
                axes[idx].set_xlabel('Time (s)')
                axes[idx].set_ylabel('Response')
            
            # Hide unused subplots
            for idx in range(n_sweeps, len(axes)):
                axes[idx].axis('off')
            
            fig.suptitle(f'Response Traces - {cellNum}', fontsize=16, fontweight='bold', y=0.995)
            plt.tight_layout()
            plt.savefig(os.path.join(out_dir, "response_grid.png"), dpi=300)
            plt.close()
            print(f"✔ Saved response grid plot → response_grid.png")
        
        # Save ALL stimulus data (mixed units)
        if stim_rows:
            sdf = pd.concat(stim_rows, ignore_index=True)
            # sdf.to_csv(os.path.join(out_dir, f"stimulus_{cellNum}.csv"), index=False)  # CSV creation disabled, using parquet only
            sdf.to_parquet(os.path.join(out_dir, f"stimulus_{cellNum}.parquet"), index=False)
            print(f"✔ Saved all stimulus → stimulus_{cellNum}.parquet")
            # Clear from memory immediately
            del sdf
        
        # Save ALL response data (mixed units)
        if resp_rows:
            rdf = pd.concat(resp_rows, ignore_index=True)
            # rdf.to_csv(os.path.join(out_dir, f"response_{cellNum}.csv"), index=False)  # CSV creation disabled, using parquet only
            rdf.to_parquet(os.path.join(out_dir, f"response_{cellNum}.parquet"), index=False)
            print(f"✔ Saved all response → response_{cellNum}.parquet")
            # Clear from memory immediately
            del rdf
        
        # Clear stimulus and response row lists from memory
        del stim_rows, resp_rows
        
        # Force garbage collection before creating unified tables
        gc.collect()
        
        # Create unified pA and mV tables (covering all sweeps with proper units)
        # Read back from saved parquet files to minimize memory usage
        if VERBOSE:
            print("\nCREATING UNIFIED UNIT TABLES")
            print("-"*70)
        
        stimulus_path = os.path.join(out_dir, f"stimulus_{cellNum}.parquet")
        response_path = os.path.join(out_dir, f"response_{cellNum}.parquet")
        
        if os.path.exists(stimulus_path) and os.path.exists(response_path):
            # Read only the columns we need for filtering
            if VERBOSE: print("  Loading stimulus data...")
            sdf = pd.read_parquet(stimulus_path, columns=['sweep', 't_s', 'value', 'unit'])
            
            if VERBOSE: print("  Creating pA table (current data)...")
            # pA table: response from VoltageClamp (0-3, 94-96) + stimulus from CurrentClamp (4-93)
            # CurrentClamp: stim is current (amperes), so use that
            stim_currentclamp = sdf[sdf['unit'].str.lower().str.contains('amp|pa', na=False)].copy()
            del sdf  # Free stimulus data immediately
            gc.collect()
            
            # Read response data
            if VERBOSE: print("  Loading response data...")
            rdf = pd.read_parquet(response_path, columns=['sweep', 't_s', 'value', 'unit'])
            
            # VoltageClamp: response is current (amperes), so use that
            resp_voltageclamp = rdf[rdf['unit'].str.lower().str.contains('amp|pa', na=False)].copy()
            
            # Combine both for pA table
            pa_table = pd.concat([stim_currentclamp, resp_voltageclamp], ignore_index=True)
            del stim_currentclamp, resp_voltageclamp  # Free intermediate data
            gc.collect()
            
            pa_table = pa_table.sort_values(['sweep', 't_s']).reset_index(drop=True)
            # pa_table.to_csv(os.path.join(out_dir, f"pA_{cellNum}.csv"), index=False)  # CSV creation disabled, using parquet only
            pa_table.to_parquet(os.path.join(out_dir, f"pA_{cellNum}.parquet"), index=False)
            print(f"✔ Saved unified pA table (all sweeps) → pA_{cellNum}.parquet")
            del pa_table
            gc.collect()
            
            # mV table: stimulus from VoltageClamp (0-3, 94-96) + response from CurrentClamp (4-93)
            if VERBOSE: print("  Creating mV table (voltage data)...")
            # Re-read stimulus for voltage filtering
            sdf = pd.read_parquet(stimulus_path, columns=['sweep', 't_s', 'value', 'unit'])
            # VoltageClamp: stim is voltage (volts), so use that
            stim_voltageclamp = sdf[sdf['unit'].str.lower().str.contains('volt|mv', na=False)].copy()
            del sdf
            gc.collect()
            
            # CurrentClamp: response is voltage (volts), so use that
            resp_currentclamp = rdf[rdf['unit'].str.lower().str.contains('volt|mv', na=False)].copy()
            del rdf  # Free response data
            gc.collect()
            
            # Combine both for mV table
            mv_table = pd.concat([stim_voltageclamp, resp_currentclamp], ignore_index=True)
            del stim_voltageclamp, resp_currentclamp
            gc.collect()
            
            mv_table = mv_table.sort_values(['sweep', 't_s']).reset_index(drop=True)
            # mv_table.to_csv(os.path.join(out_dir, f"mV_{cellNum}.csv"), index=False)  # CSV creation disabled, using parquet only
            mv_table.to_parquet(os.path.join(out_dir, f"mV_{cellNum}.parquet"), index=False)
            print(f"✔ Saved unified mV table (all sweeps) → mV_{cellNum}.parquet")
            del mv_table
            gc.collect()

        if stim_plot_data is not None:
            del stim_plot_data
        if resp_plot_data is not None:
            del resp_plot_data
        
        # Get sampling rates - detect if mixed
        sampling_rates = set()
        if stim:
            for n, c in stim.values():
                if hasattr(c, 'rate') and c.rate:
                    sampling_rates.add(float(c.rate))
        if acq:
            for n, c in acq.values():
                if hasattr(c, 'rate') and c.rate:
                    sampling_rates.add(float(c.rate))
        
        print(f"\n{'='*70}")
        print(f"✅ EXTRACTION COMPLETE")
        print(f"{'='*70}\n")
        
        # Close any remaining matplotlib figures to free memory
        plt.close('all')
        
        # Return sampling rates and protocol info for manifest creation in main loop
        return sampling_rates, protocol_info


def process_and_log_mixed_protocol_files(parent_dir, log_output_dir, output_filename="mixed_protocol_log.xlsx", template_path=None, cell_id_prefix="", specific_nwb_file=None, cell_count_override=None):
    """
    Process multiple mixed protocol NWB files and create metadata log Excel file using template structure.
    
    Args:
        parent_dir: Directory containing NWB files or subdirectories with NWB files
        log_output_dir: Directory where Excel log will be saved
        output_filename: Name of the output Excel file (default: mixed_protocol_log.xlsx)
        template_path: Path to Excel template file (required)
        cell_id_prefix: Optional prefix for cell IDs
        specific_nwb_file: If provided, only process this specific NWB file (command-line mode)
        cell_count_override: If provided (from main.py), use this cell count instead of auto-incrementing
    """
    os.makedirs(log_output_dir, exist_ok=True)
    
    # Load Excel template if provided
    column_names = []
    if template_path and Path(template_path).is_file():
        try:
            wb = load_workbook(template_path)
            ws = wb.active
            HEADER_ROW = 3
            for col in range(1, ws.max_column + 1):
                val = ws.cell(row=HEADER_ROW, column=col).value
                if val is None:
                    break
                column_names.append(val)
            if VERBOSE: print(f"✓ Loaded template with {len(column_names)} columns")
        except Exception as e:
            print(f"⚠ Could not load template: {e}")
            column_names = []
    
    # Add computed columns if not already in template
    if "brain_loc" not in column_names:
        column_names.append("brain_loc")
    # Note: sampleRate_Hz is no longer added to the Excel
    
    all_rows = []
    
    # Search for NWB files
    nwb_files = []
    
    if specific_nwb_file:
        # Command-line mode: process only the specified file
        nwb_files = [Path(specific_nwb_file)]
        print(f"\n→ Processing single specified file: {Path(specific_nwb_file).name}")
    else:
        # Interactive mode: scan directory for all files
        parent_path = Path(parent_dir)
        
        # Look in parent directory and subdirectories
        if parent_path.is_file() and parent_path.suffix == '.nwb':
            nwb_files = [parent_path]
        else:
            nwb_files = list(parent_path.glob('**/*.nwb'))
        
        print(f"\nFound {len(nwb_files)} NWB files to process")
    
    # Track cell count per subject_id
    subject_cell_counts = {}

    
    for idx, nwb_path in enumerate(nwb_files, 1):
        # Only show counter if processing multiple files
        if len(nwb_files) > 1:
            print(f"\n[{idx}/{len(nwb_files)}] Processing: {nwb_path.name}")
        
        try:
            with NWBHDF5IO(str(nwb_path), 'r') as io:
                nwb = io.read()
                
                # Check if this is a mixed protocol file (only needed in interactive mode)
                if not specific_nwb_file:
                    # Interactive mode: check and skip single protocol files
                    protocols = set()
                    if nwb.stimulus:
                        for stim_name in nwb.stimulus.keys():
                            stim_series = nwb.stimulus[stim_name]
                            protocol = get_protocol_type(stim_series)
                            protocols.add(protocol)
                    
                    # Skip if not mixed protocol (only process files with 2+ protocol types)
                    if len(protocols) < 2:
                        print(f"  ⊘ Skipping: Single protocol file (not mixed)")
                        continue
                # In command-line mode, main.py already verified it's mixed protocol, so skip check
                
                # Extract metadata
                cell_id = cell_id_prefix or nwb_path.stem.split('_')[0]
                session_start = nwb.session_start_time
                nwb_id = safe_getattr(nwb, 'identifier', '')
                
                # Extract session info
                session_desc = safe_getattr(nwb, 'session_description', '')
                experiment_desc = safe_getattr(nwb, 'experiment_description', '')
                
                # Extract subject info
                subject = nwb.subject
                subject_id = safe_getattr(subject, 'subject_id', '')
                age = safe_getattr(subject, 'age', '')
                age = normalize_age(age)  # Convert days to years if needed
                sex = safe_getattr(subject, 'sex', '')
                genotype = safe_getattr(subject, 'genotype', '')
                species = safe_getattr(subject, 'species', '')
                
                # Track cell count per subject
                # If count was pre-initialized (command-line mode), use it; otherwise auto-increment (interactive mode)
                if cell_count_override is not None:
                    # Command-line mode: use the count passed from main.py
                    cell_count = cell_count_override
                else:
                    # Interactive mode: auto-increment per subject
                    if subject_id not in subject_cell_counts:
                        subject_cell_counts[subject_id] = 0
                    subject_cell_counts[subject_id] += 1
                    cell_count = subject_cell_counts[subject_id]
                
                # Extract lab info
                lab = safe_getattr(nwb, 'lab', '')
                institution = safe_getattr(nwb, 'institution', '')
                experimenter = safe_getattr(nwb, 'experimenter', [])
                
                # Use institution as fallback for facility if lab is not available
                facility = lab if lab else institution
                
                # Extract keywords for parsing - handle None case
                keywords_raw = safe_getattr(nwb, 'keywords', None)
                if keywords_raw is None:
                    keywords = []
                else:
                    try:
                        keywords = list(keywords_raw)
                    except TypeError:
                        keywords = [keywords_raw] if keywords_raw else []
                
                # Count sweeps and protocols - with proper None checks
                if not nwb.stimulus:
                    raise ValueError("No stimulus data found in NWB file")
                
                num_sweeps = len(list(nwb.stimulus.keys()))
                protocols = set()
                sampling_rates = set()
                protocol_info = {}  # Track protocol and rate per sweep
                
                for stim_name in nwb.stimulus.keys():
                    stim_series = nwb.stimulus[stim_name]
                    protocol = get_protocol_type(stim_series)
                    protocols.add(protocol)
                    rate = getattr(stim_series, 'rate', 50000.0)
                    sampling_rates.add(rate)
                    # Extract sweep number and store protocol info
                    sweep_num = extract_sweep_number(stim_name)
                    protocol_info[sweep_num] = {'protocol': protocol, 'rate': rate}
                
                # Map extracted metadata to template columns
                row = {col: "" for col in column_names}  # Initialize with empty strings
                
                # Map metadata to template columns (following single protocol structure)
                if "recDate" in row:
                    row["recDate"] = session_start.strftime("%Y%m%d") if session_start else ""
                if "facility" in row:
                    row["facility"] = facility
                if "genotype" in row:
                    row["genotype"] = genotype
                if "sex" in row:
                    row["sex"] = sex
                if "age_at_recDate" in row:
                    row["age_at_recDate"] = age
                if "recGoal" in row:
                    # For mixed protocol: use experiment_desc if available, otherwise describe protocols
                    if experiment_desc:
                        row["recGoal"] = experiment_desc[:50]
                    else:
                        # Create descriptive rec_goal from protocol types
                        protocol_list = sorted(list(protocols))
                        row["recGoal"] = f"Mixed protocol: {', '.join(protocol_list)}"
                if "cellNum" in row:
                    row["cellNum"] = f"{subject_id}_{cell_count}"
                if "fileNum" in row:
                    row["fileNum"] = nwb_id
                if "expInfoTime" in row:
                    row["expInfoTime"] = session_start.isoformat() if session_start else ""
                if "cellDepth" in row:
                    row["cellDepth"] = parse_cell_depth(keywords)
                if "cellType" in row:
                    row["cellType"] = parse_cell_type(keywords)
                if "remarks" in row:
                    row["remarks"] = subject_id
                if "brain_loc" in row:
                    # First try to parse from keywords
                    brain_loc = "Human Neocortex" if any("Human Neocortex" in k for k in keywords) else ""
                    # If not found in keywords, check species and institution
                    if not brain_loc:
                        is_human = species and "homo sapiens" in species.lower()
                        is_allen = institution and "allen institute" in institution.lower()
                        if is_human and is_allen:
                            brain_loc = "Human Neocortex"
                    row["brain_loc"] = brain_loc
                
                # Add sampling rate to row metadata
                if len(sampling_rates) == 1:
                    # Single sampling rate - store as scalar
                    row['sampleRate_Hz'] = list(sampling_rates)[0]
                elif len(sampling_rates) > 1:
                    # Multiple sampling rates - store as list
                    print(f"  ⚠ WARNING: Found mixed sampling rates: {sampling_rates}")
                    row['sampleRate_Hz'] = sorted(list(sampling_rates), reverse=True)
                else:
                    # Fallback if no rates detected
                    row['sampleRate_Hz'] = 50000.0
                # Note: sampleRate_Hz is no longer added to Excel, but is in manifest
                
                all_rows.append(row)
                print(f"  ✓ Extracted metadata: {num_sweeps} sweeps, {len(protocols)} protocol types")
                
                # Create output folder and extract data to CSV/Parquet files with correct cellNum
                cell_folder = nwb_path.parent / nwb_path.stem
                cell_folder.mkdir(parents=True, exist_ok=True)  # Create folder if it doesn't exist
                
                print(f"  🔄 Extracting data files with cellNum: {row['cellNum']}")
                try:
                    extract_from_mixed_protocol_nwb(
                        str(nwb_path), 
                        str(cell_folder), 
                        row['cellNum'], 
                        plot=True
                    )
                except Exception as extract_err:
                    print(f"  ⚠ Warning: Data extraction failed: {extract_err}")
                    import traceback
                    traceback.print_exc()
                
                # Create manifest for this cell
                manifest_path = cell_folder / "manifest.json"
                manifest = {
                    "file_id": nwb_id,
                    "nwb_path": str(nwb_path.absolute()),
                    "tables": {
                        "pa": f"pA_{row['cellNum']}.parquet",
                        "mv": f"mV_{row['cellNum']}.parquet",
                        "stimulus": f"stimulus_{row['cellNum']}.parquet",
                        "response": f"response_{row['cellNum']}.parquet"
                    },
                    "meta": row,
                    "protocols": protocol_info
                }
                
                with open(manifest_path, "w") as f:
                    json.dump(manifest, f, indent=2, default=str)
                print(f"  ✓ Saved manifest → {manifest_path}")
                
        except Exception as e:
            print(f"  ✗ Error processing {nwb_path.name}: {e}")
            import traceback
            traceback.print_exc()  # Print full stack trace for debugging
            continue
        finally:
            # Force garbage collection after each file to free memory
            # This is critical for processing many large NWB files
            gc.collect()
    
    # Save to Excel (append if exists, create if not)
    if all_rows:
        log_path = Path(log_output_dir) / output_filename
        
        # Check if file exists
        if log_path.exists():
            # Append to existing Excel file
            existing_df = pd.read_excel(log_path)
            new_df = pd.DataFrame(all_rows, columns=column_names)
            
            # Ensure columns match
            for col in new_df.columns:
                if col not in existing_df.columns:
                    existing_df[col] = ""
            for col in existing_df.columns:
                if col not in new_df.columns:
                    new_df[col] = ""
            
            # Append new rows
            combined_df = pd.concat([existing_df, new_df], ignore_index=True)
            combined_df.to_excel(log_path, index=False)
            print(f"\n✔ Appended to existing metadata log → {log_path}")
            print(f"  New rows added: {len(all_rows)}")
            print(f"  Total rows now: {len(combined_df)}")
        else:
            # Create new Excel file
            df = pd.DataFrame(all_rows, columns=column_names)
            df.to_excel(log_path, index=False)
            print(f"\n✔ Created new metadata log → {log_path}")
            print(f"  Total rows: {len(df)}")
        
        if VERBOSE: print(f"  Columns: {len(column_names)}")
    else:
        print("⚠ No NWB files were successfully processed")


if __name__ == "__main__":
    import sys
    
    # Support command-line arguments from main.py
    if len(sys.argv) > 3:
        # Called from main.py with arguments: parent_dir, log_output_dir, template_path, nwb_file, [cell_count]
        if VERBOSE: print("→ Using command-line arguments (called from main.py)")
        parent_dir = Path(sys.argv[1]).expanduser()
        log_output_dir = Path(sys.argv[2]).expanduser()
        template_path_input = Path(sys.argv[3]).expanduser()
        nwb_file_arg = sys.argv[4] if len(sys.argv) > 4 else None
        cell_count_arg = int(sys.argv[5]) if len(sys.argv) > 5 else None
        
        # Template validation already done in main.py - just use the path
        if VERBOSE:
            print(f"  parent_dir = {parent_dir}")
            print(f"  log_output_dir = {log_output_dir}")
            print(f"  template_path = {template_path_input}")
            print(f"  nwb_file = {nwb_file_arg}")
            print(f"  cell_count = {cell_count_arg}\n")
    else:
        # Interactive mode (standalone use)
        if VERBOSE: print("→ Using interactive mode (prompting for input)")
        parent_dir = Path(input("Parent directory for all subjects: ")).expanduser()
        log_output_dir = Path(input("Output directory for excel metadata: ")).expanduser()
        template_path_input = Path(input("Path to template Excel metadata: ")).expanduser()
        nwb_file_arg = None
        cell_count_arg = None
    
    # Validate parent directory
    if not parent_dir.is_dir():
        raise FileNotFoundError(f"Parent directory not found: {parent_dir}")
    
    # Prepare output directory
    log_output_dir.mkdir(parents=True, exist_ok=True)
    
    # Validate template exists
    if not template_path_input.is_file():
        raise FileNotFoundError(f"Template Excel file not found: {template_path_input}")
    
    # Support both interactive and command-line modes
    if nwb_file_arg:
        # Command-line mode (called from main.py with NWB file)
        nwb = nwb_file_arg
        
        # Auto-generate output directory
        nwb_path = Path(nwb)
        cell_id = nwb_path.stem.split('_')[0]  # Extract subject ID from filename
        out = str(nwb_path.parent / nwb_path.stem)
        
        plt_on = True  # Generate plots to match single protocol behavior
        
        # Data extraction now happens inside process_and_log_mixed_protocol_files
        # which has access to the correct cellNum format
        
        # Log to Excel and extract data - ONLY process the specified NWB file
        process_and_log_mixed_protocol_files(
            parent_dir, 
            str(log_output_dir), 
            output_filename="human_ephys_log.xlsx", 
            template_path=str(template_path_input), 
            cell_id_prefix=cell_id,
            specific_nwb_file=nwb_file_arg,  # Only process this file
            cell_count_override=cell_count_arg  # Use cell count passed from main.py
        )
    else:
        # Interactive mode (standalone use)
        nwb = input("NWB path: ").strip()
        
        # Auto-generate output directory
        nwb_path = Path(nwb)
        cell_id = nwb_path.stem.split('_')[0]  # Extract subject ID from filename
        default_out = str(nwb_path.parent / nwb_path.stem)
        
        out = input(f"Output dir (default: {default_out}): ").strip() or default_out
        
        plt_on = input("Generate plots? (yes): ").lower() in ['yes', 'y']
        
        # Data extraction now happens inside process_and_log_mixed_protocol_files
        # which has access to the correct cellNum format
        
        # Log to Excel and extract data
        process_and_log_mixed_protocol_files(parent_dir, str(log_output_dir), output_filename="human_ephys_log.xlsx", template_path=str(template_path_input), cell_id_prefix=cell_id)
